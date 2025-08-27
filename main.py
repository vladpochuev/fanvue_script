import math
import os
import sys
import threading
import time
import tkinter as tk
from tkinter import filedialog, ttk

import ssl

import undetected_chromedriver as uc
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

ssl._create_default_https_context = ssl._create_unverified_context

def resource_path(relative_path):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def choose_file():
    file_path = filedialog.askopenfilename(
        title="Select excel file",
        filetypes=(("Excel files", "*.xlsx *.xls"),)
    )
    if file_path:
        update_status(f"Selected Excel File: {file_path}")
        threading.Thread(target=start_script, args=(file_path,), daemon=True).start()


def update_progress_bar(idx, quantity):
    percent = math.floor(idx / quantity * 100)
    root.after(0, lambda: (
        progress_bar.config(value=percent),
        lbl_progress.config(text=f"Processed {idx} of {quantity} accounts ({percent}%)")
    ))


def update_status(text):
    root.after(0, lambda: lbl_status.config(text=text))


def start_script(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    accounts = []
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        email = row[0]
        password = row[1]
        if email and password:
            accounts.append((email, password))

    update_progress_bar(0, len(accounts))

    for idx, account in enumerate(accounts, start=1):
        update_status(f"Opening browser for account {idx}/{len(accounts)}")

        chrome_path = resource_path("chrome/GoogleChromePortable/App/Chrome-bin/chrome.exe")

        options = uc.ChromeOptions()
        options.add_argument("--start-maximized")
        options.binary_location = chrome_path
        driver = uc.Chrome(options=options)

        try:
            update_status("Opening login page...")
            driver.get("https://www.fanvue.com/signin")

            email_input = driver.find_element(By.NAME, "email")
            email_input.send_keys(account[0])

            password_input = driver.find_element(By.NAME, "password")
            password_input.send_keys(account[1])

            WebDriverWait(driver, timeout=10 ** 9).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']"))
            )

            update_status("Logging in...")
            driver.find_element(By.CSS_SELECTOR, 'button[type="submit"]').click()
            WebDriverWait(driver, 10).until(EC.url_to_be("https://www.fanvue.com/home"))

            update_status("Opening creator page...")
            driver.get("https://www.fanvue.com/daisy_grace_uk")
            time.sleep(5)

            try:
                driver.find_element(By.CSS_SELECTOR,
                                    'button[data-sentry-source-file="Over18ConsentDialog.tsx"]').click()
            except:
                pass

            follow = driver.find_element(By.CSS_SELECTOR, 'button.mui-5h07u6')
            try:
                follow_span = follow.find_element(By.CSS_SELECTOR, 'span')
                if follow_span.text == "Follow for free":
                    update_status("Following creator...")
                    follow.click()

                WebDriverWait(driver, timeout=10).until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, 'button[data-sentry-source-file="UpgradeToSubscriptionDialog.tsx"]')))
                driver.find_element(By.CSS_SELECTOR,
                                    'button[data-sentry-source-file="UpgradeToSubscriptionDialog.tsx"]').click()
            except:
                pass

            update_status("Collecting posts...")
            posts = driver.find_elements(By.CSS_SELECTOR, 'div[aria-label="post"]')
            resolved = []

            while posts:
                for post in posts:
                    resolved.append(post)
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", post)
                    try:
                        like = post.find_element(By.CSS_SELECTOR, 'button[aria-label="Unlike post"]')
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", like)
                        like.click()
                    except:
                        pass
                    time.sleep(1)

                try:
                    driver.execute_script(
                        "document.querySelector('header[data-sentry-element=\"PostHeader\"]').scrollIntoView({block: 'center'});")
                except:
                    pass

                time.sleep(7)
                new_posts = driver.find_elements(By.CSS_SELECTOR, 'div[aria-label="post"]')
                posts = [x for x in new_posts if x not in resolved]

            update_status(f"Account {idx}/{len(accounts)} finished")

        finally:
            driver.quit()
            update_progress_bar(idx, len(accounts))


root = tk.Tk()
btn_choose = tk.Button(root, text="Choose excel file", command=choose_file)
lbl_status = tk.Label(root, text="File is not selected")
progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
lbl_progress = tk.Label(root, text="")


def run():
    root.title("Fanvue autolike script")
    root.geometry("500x220")

    btn_choose.pack(pady=10)
    lbl_status.pack(pady=10)
    progress_bar.pack(pady=10)
    lbl_progress.pack(pady=5)

    root.mainloop()


if __name__ == "__main__":
    run()
